import pandas as pd
import numpy as np

CANAL_CUESTA    = '007 - TD_MAQUILA'
CANAL_FRESCAMPO = '004 - GS_MAQUILA ÉXITO'
CANAL_LATTI     = '007 - TD_MAQUILA_D1'
CANAL_PUNTOS    = '005 - PUNTOS DE VENTAS'
CANAL_OTROS     = '008 - OTROS CANALES'

CO_PUNTOS    = {2, 3, 4, 5, 6}
CO_PRINCIPAL = 1

COLUMN_MAP = {
    'Fecha': 'fecha',
    'C.O.': 'co',
    'Desc. C.O.': 'desc_co',
    'Cliente factura': 'cliente_factura',
    'Razon social cliente factura': 'razon_social',
    'Sucursal factura': 'sucursal_factura',
    'Lista de precios': 'lista_precios',
    'CATEGORIA': 'categoria',
    'CLASES DE CLIENTES': 'clases_clientes',
    'Tipo docto.': 'tipo_docto',
    'Nro documento': 'nro_documento',
    'Item': 'item_num',
    'Referencia': 'referencia',
    'Desc. item': 'desc_item',
    'Desc. tipo inventario': 'desc_tipo_inventario',
    'U.M. inv.': 'um_inv',
    'Desc. motivo': 'desc_motivo',
    'Vendedor': 'vendedor',
    'Nombre vendedor': 'nombre_vendedor',
    'Desc. ciudad': 'desc_ciudad',
    'Cantidad': 'cantidad',
    'Desc. U.N.': 'desc_un',
    'Vol?men en LTR ': 'volumen_ltr',
    'Volumen en LTR ': 'volumen_ltr',
    'Precio unit.': 'precio_unit',
    'Valor descuentos': 'valor_descuentos',
    'Descuento 1': 'descuento_1',
    'Descuento 2': 'descuento_2',
    'Descuento 3': 'descuento_3',
    'Dscto. promedio %': 'dscto_promedio_pct',
    'Valor impuestos': 'valor_impuestos',
    'Vlr. imp. IBUA': 'vlr_imp_ibua',
    'Vlr. imp. ICUI': 'vlr_imp_icui',
    'Vlr. imp. IVA': 'vlr_imp_iva',
    'Valor neto': 'valor_neto',
    'Valor bruto': 'valor_bruto',
    'Valor subtotal': 'valor_subtotal',
    'Costo promedio total': 'costo_promedio_total',
    'Margen promedio': 'margen_promedio',
    'CANAL DE VENTAS': 'canal_ventas',
    'RUTAS DE VENTAS': 'rutas_ventas',
    'SUB-GRUPO': 'sub_grupo',
    'ZONAS': 'zonas',
    'GRUPOS': 'grupos',
}

CURRENCY_COLS = [
    'precio_unit', 'valor_descuentos', 'valor_impuestos',
    'vlr_imp_ibua', 'vlr_imp_icui', 'vlr_imp_iva',
    'valor_neto', 'valor_bruto', 'valor_subtotal', 'costo_promedio_total',
]

MONTH_MAP = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE',
}

FORMULA_COLS = [
    'UM.LITRO_SUERO', 'FAMILIA', 'TIPO DE LECHE ', 'TIPO DE LECHE',
    'LITROS_LECHE', 'KILOS_SUERO', 'RENTABILIDAD',
    'RENTABILIDAD PLATA', 'ITEM+DESCRIPCION', 'MES',
]


def read_file(file_obj, filename: str) -> pd.DataFrame:
    ext = filename.rsplit('.', 1)[-1].lower()

    if ext in ('txt', 'csv'):
        for enc in ('latin-1', 'cp1252', 'utf-8-sig', 'utf-8'):
            try:
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)
                df = pd.read_csv(
                    file_obj, sep='\t',
                    encoding=enc, low_memory=False, dtype=str,
                    on_bad_lines='skip',
                )
                # Drop unnamed index column if present (some exports add row numbers)
                if df.columns[0].strip() in ('', 'Unnamed: 0') or df.columns[0].strip().isdigit():
                    df = df.drop(columns=df.columns[0])
                # Drop trailing empty columns
                unnamed = [c for c in df.columns if str(c).startswith('Unnamed:')]
                df = df.drop(columns=unnamed, errors='ignore')
                df = df.reset_index(drop=True)
                return df
            except Exception:
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)
        raise ValueError("No se pudo leer el archivo TXT")

    elif ext in ('xlsx', 'xls'):
        try:
            df = pd.read_excel(file_obj, sheet_name='DATA', dtype=str)
        except Exception:
            df = pd.read_excel(file_obj, dtype=str)
        drop = [c for c in FORMULA_COLS if c in df.columns]
        df = df.drop(columns=drop, errors='ignore')
        return df

    raise ValueError(f"Formato no soportado: .{ext}")


def _clean_currency(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace(r'[\$,]', '', regex=True)
        .str.strip()
        .replace('', '0')
        .pipe(pd.to_numeric, errors='coerce')
        .fillna(0)
    )


def process(df: pd.DataFrame, items_df: pd.DataFrame) -> tuple:
    stats = {}
    initial = len(df)

    # Rename columns to DB names
    df = df.rename(columns={k: v for k, v in COLUMN_MAP.items() if k in df.columns})

    required = ['canal_ventas', 'valor_subtotal', 'desc_item', 'co']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Columnas faltantes: {missing}")

    # Clean currency
    for col in CURRENCY_COLS:
        if col in df.columns:
            df[col] = _clean_currency(df[col])

    # Clean plain numerics
    for col in ['cantidad', 'volumen_ltr', 'dscto_promedio_pct', 'margen_promedio']:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(',', '', regex=False),
                errors='coerce',
            ).fillna(0)

    df['co'] = pd.to_numeric(df['co'], errors='coerce').fillna(0).astype(int)
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')

    # Rule 1 – remove subtotal = 0
    df = df[df['valor_subtotal'] != 0].copy()
    stats['eliminadas_subtotal_cero'] = initial - len(df)

    desc = df['desc_item'].fillna('')

    # Rule 2 – cuesta
    m2 = desc.str.contains('cuesta', case=False, na=False)
    df.loc[m2, 'canal_ventas'] = CANAL_CUESTA
    stats['reclasificadas_cuesta'] = int(m2.sum())

    # Rule 3 – Frescampo
    m3 = desc.str.contains('Frescampo', case=False, na=False)
    df.loc[m3, 'canal_ventas'] = CANAL_FRESCAMPO
    stats['reclasificadas_frescampo'] = int(m3.sum())

    # Rule 4 – Latti
    m4 = desc.str.contains('Latti', case=False, na=False)
    df.loc[m4, 'canal_ventas'] = CANAL_LATTI
    stats['reclasificadas_latti'] = int(m4.sum())

    # Rule 5 – CO 02-06 not puntos → puntos
    m5 = df['co'].isin(CO_PUNTOS) & (df['canal_ventas'] != CANAL_PUNTOS)
    df.loc[m5, 'canal_ventas'] = CANAL_PUNTOS
    stats['reclasificadas_a_puntos'] = int(m5.sum())

    # Rule 6 – CO=01 puntos → otros
    m6 = (df['co'] == CO_PRINCIPAL) & (df['canal_ventas'] == CANAL_PUNTOS)
    df.loc[m6, 'canal_ventas'] = CANAL_OTROS
    stats['reclasificadas_a_otros'] = int(m6.sum())

    # Formula columns (VLOOKUP equivalents)
    ref = df['referencia'].astype(str).str.strip()

    tipo_map, peso_map, leche_map = {}, {}, {}
    if not items_df.empty and 'referencia' in items_df.columns:
        items = items_df.copy()
        items['referencia'] = items['referencia'].astype(str).str.strip()
        items = items.drop_duplicates(subset='referencia').set_index('referencia')
        tipo_map  = items['tipo'].to_dict()       if 'tipo'       in items.columns else {}
        peso_map  = items['peso'].to_dict()       if 'peso'       in items.columns else {}
        leche_map = items['tipo_leche'].to_dict() if 'tipo_leche' in items.columns else {}

    df['familia']        = ref.map(tipo_map)
    df['um_litro_suero'] = pd.to_numeric(ref.map(peso_map), errors='coerce').fillna(0)
    df['tipo_leche']     = ref.map(leche_map)

    df['litros_leche']      = df['cantidad'] * df['um_litro_suero']
    df['kilos_suero']       = df['cantidad'] * df['um_litro_suero']
    subtotal                = df['valor_subtotal']
    costo                   = df['costo_promedio_total']
    df['rentabilidad']      = np.where(subtotal != 0, (subtotal - costo) / subtotal, 0.0)
    df['rentabilidad_plata']= subtotal - costo
    df['item_descripcion']  = ref + ' ' + df['desc_item'].fillna('')

    # Mes/año from date
    df['mes']  = df['fecha'].dt.month.map(MONTH_MAP)
    df['anio'] = df['fecha'].dt.year.astype('Int64')

    str_cols = df.select_dtypes(include='object').columns
    df[str_cols] = df[str_cols].fillna('')

    stats['filas_finales'] = len(df)
    return df, stats


HAPPY_CO_SET = {2, 3, 4, 5, 6}


def read_horas_from_excel(file_obj) -> pd.DataFrame:
    """
    Lee el archivo 'Ventas con hora' (cabecera de factura con hora de aprobación).
    Devuelve: nro_documento, co, fecha, hora, hora_num, dia_semana, es_happy.
    es_happy = miércoles + 14:00–15:00 + CO en 002-006.
    Tolerante a variaciones en los headers (tildes, mayúsculas, espacios).
    """
    import openpyxl, datetime, re, unicodedata

    def _norm(s):
        # Sin tildes, minúsculas, sin espacios/puntos extras
        if s is None:
            return ''
        s = unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode()
        return re.sub(r'[^a-z0-9]', '', s.lower())

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    # mapa normalizado: 'horaaprobacion' → índice
    hmap = {_norm(h): i for i, h in enumerate(header) if h is not None}

    i_doc   = hmap.get('nrodocumento')
    i_co    = hmap.get('co')
    i_fecha = hmap.get('fecha')
    i_hora  = hmap.get('horaaprobacion')

    if i_doc is None or i_hora is None:
        raise ValueError(
            "El archivo de horas debe tener columnas 'Nro documento' y "
            "'Hora aprobacion' (se aceptan con o sin tildes). "
            f"Columnas encontradas: {[h for h in header if h]}"
        )

    _re_hora = re.compile(r'(\d+):(\d+):(\d+)\s*(AM|PM)?')
    out = []
    for r in it:
        if r[i_doc] is None:
            continue
        doc = str(r[i_doc]).strip()
        if not doc:
            continue

        try:
            co = int(str(r[i_co]).strip()) if i_co is not None and r[i_co] is not None else 0
        except (ValueError, TypeError):
            co = 0

        fecha = r[i_fecha] if i_fecha is not None else None
        dia = fecha.weekday() if isinstance(fecha, datetime.datetime) else None
        fecha_str = fecha.strftime('%Y-%m-%d') if isinstance(fecha, datetime.datetime) else None

        hora_s = str(r[i_hora]).strip() if r[i_hora] is not None else ''
        hora_num = None
        m = _re_hora.match(hora_s.upper())
        if m:
            hh, mm = int(m.group(1)), int(m.group(2))
            ap = m.group(4)
            if ap == 'PM' and hh != 12:
                hh += 12
            elif ap == 'AM' and hh == 12:
                hh = 0
            hora_num = round(hh + mm / 60.0, 3)

        es_happy = bool(
            dia == 2 and hora_num is not None
            and 14.0 <= hora_num < 15.0 and co in HAPPY_CO_SET
        )

        out.append({
            'nro_documento': doc, 'co': co, 'fecha': fecha_str,
            'hora': hora_s, 'hora_num': hora_num,
            'dia_semana': dia, 'es_happy': es_happy,
        })

    return pd.DataFrame(out)


def read_items_from_excel(file_obj) -> pd.DataFrame:
    # Intenta hoja ITEM primero; si no existe usa la primera hoja
    try:
        df = pd.read_excel(file_obj, sheet_name='ITEM', dtype=str)
    except Exception:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=0, dtype=str)

    # Normalizar nombres de columna
    df.columns = [str(c).strip() for c in df.columns]

    expected = ['Referencia Item', 'Nombre Item', 'Tipo', 'PESO', 'VOLUMEN', 'TIPO DE LECHE']
    if df.columns[0] in expected or df.shape[1] >= 6:
        df = df.iloc[:, :6]
        df.columns = ['referencia', 'nombre_item', 'tipo', 'peso', 'volumen', 'tipo_leche']
    else:
        # Renombrar las primeras 6 columnas a nombres estándar
        cols = list(df.columns) + [''] * 6
        df = df.rename(columns={
            cols[0]: 'referencia', cols[1]: 'nombre_item',
            cols[2]: 'tipo',       cols[3]: 'peso',
            cols[4]: 'volumen',    cols[5]: 'tipo_leche',
        })

    df['referencia'] = df['referencia'].astype(str).str.strip()
    df['peso']       = pd.to_numeric(df['peso'],    errors='coerce').fillna(0)
    df['volumen']    = pd.to_numeric(df['volumen'], errors='coerce').fillna(0)

    # Eliminar filas sin referencia válida
    df = df[~df['referencia'].isin(['', 'nan', 'None', 'NaN'])]
    return df.dropna(subset=['referencia'])
